#!/usr/bin/env python3
"""Convertit une livraison de cartes en JSON de `assets/decks/`.

Outil d'atelier : il tourne chez nous, jamais sur le téléphone. Il lit un CSV
exporté d'une feuille de calcul, tolère ce qu'une feuille produit réellement —
séparateur local, BOM, en-têtes accentués, colonnes de travail — et écrit le
format que le seeder attend.

**Il n'ignore jamais une ligne inexploitable.** Il refuse d'écrire quoi que ce
soit et liste toutes les fautes d'un coup : une carte perdue en silence est
invisible jusqu'à ce qu'un joueur la cherche dans le jeu, et corriger un
fichier de six cents lignes une faute à la fois est intenable.

**Il ne normalise pas les textes.** Le dédoublonnage de R6.4 se fait sur un
texte normalisé — casse, accents, ligatures, élisions, ponctuation — et cette
normalisation vit dans le moteur, en Dart, testée. La réimplémenter ici en
ferait deux versions qui divergeraient. Ce script n'attrape que les doublons
franchement identiques ; le reste est attrapé en CI par un test Dart qui relit
tous les JSON avec la vraie fonction du moteur.

Usage :

    python tool/import_decks.py cartes.csv \\
        --out assets/decks/animaux.json \\
        --id animaux --name Animaux --audience family --min-age 6

Si le JSON de sortie existe déjà, ses métadonnées sont reprises pour tout ce
qui n'est pas donné en argument, et `contentVersion` est incrémenté — c'est lui
qui déclenche le re-seeding des installations existantes.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import sys
import unicodedata
from pathlib import Path

#: Longueur au-delà de laquelle une carte devient illisible à bout de bras.
#:
#: `docs/CONTENU.md` vise « 30 caractères environ » pour un mot à faire
#: deviner, et c'est toujours la cible. Mais la livraison d'août 2026 a
#: apporté des catégories de **situations** — « Se cogner le petit orteil dans
#: le meuble » — dont le format est intrinsèquement plus long qu'un nom commun.
#: La limite passe donc de 40 à 60 : assez pour une situation, trop peu pour
#: un paragraphe. Ce qu'elle protège reste la lisibilité d'un coup d'œil, pas
#: une norme d'écriture.
MAX_TEXT_LENGTH = 60

#: `docs/CONTENU.md` : en hésitation entre 1 et 2, choisir 2.
DEFAULT_DIFFICULTY = 2

#: Niveaux écrits en toutes lettres, tels qu'une feuille les porte réellement.
#:
#: La livraison d'août 2026 note « 🟢 Facile », « 🟡 Moyen », « 🔴 Expert » :
#: une pastille de couleur pour trier à l'œil dans le tableur, et un mot. Les
#: convertir à la main avant l'import remettrait un humain sur le chemin de
#: 500 lignes, ce que cet outil existe précisément pour éviter.
DIFFICULTY_ALIASES = {
    "facile": 1,
    "simple": 1,
    "moyen": 2,
    "moyenne": 2,
    "intermediaire": 2,
    "difficile": 3,
    "expert": 3,
    "dur": 3,
}

#: Séparateurs qu'une feuille de calcul produit selon sa locale.
CANDIDATE_DELIMITERS = [",", ";", "\t"]

#: En-têtes acceptés, après mise à plat, pour chaque champ.
HEADER_ALIASES = {
    "text": {"texte", "text", "carte", "mot"},
    "difficulty": {"difficulte", "difficulty", "niveau"},
    "category": {"categorie", "category", "theme", "thematique", "deck"},
}


class ImportError_(Exception):
    """Livraison inexploitable. Le message liste toutes les fautes."""


def _flatten(value: str) -> str:
    """Met un en-tête à plat : sans accent, sans espace, en minuscules.

    Ne sert **qu'aux en-têtes**. Le texte des cartes n'est jamais transformé :
    c'est du contenu, et « Éléphant » doit rester « Éléphant ».
    """
    sans_accent = "".join(
        c
        for c in unicodedata.normalize("NFD", value)
        if unicodedata.category(c) != "Mn"
    )
    return sans_accent.strip().lower().replace(" ", "").replace("_", "")


def _clean_cell(value: str | None) -> str:
    """Nettoie une cellule sans toucher au sens du texte.

    Les espaces insécables et fines viennent des copier-coller depuis un
    traitement de texte ; ils sont invisibles à l'œil et casseraient les
    comparaisons de doublons.
    """
    if value is None:
        return ""
    espaces = {" ": " ", " ": " ", " ": " ", "﻿": ""}
    for avant, apres in espaces.items():
        value = value.replace(avant, apres)
    return " ".join(value.split())


def _sniff_delimiter(header_line: str) -> str:
    """Le séparateur le plus présent dans la ligne d'en-tête."""
    return max(CANDIDATE_DELIMITERS, key=header_line.count)


def _map_columns(fieldnames: list[str]) -> dict[str, str]:
    """Associe chaque champ attendu au nom de colonne réellement livré."""
    mapping: dict[str, str] = {}
    for champ, alias in HEADER_ALIASES.items():
        for nom in fieldnames:
            if _flatten(nom or "") in alias:
                mapping[champ] = nom
                break
    return mapping


def slugify(value: str) -> str:
    """Identifiant de deck dérivé d'un nom de catégorie.

    Ne sert **qu'aux identifiants de deck**, jamais aux cartes : celles-ci
    reçoivent leur identifiant du seeder, en Dart, sous la forme
    `<deckId>:<slug du texte>`. Aucun risque de divergence avec la
    normalisation de R6.4, qui vit ailleurs et ne passe pas par ici.
    """
    plat = "".join(
        c
        for c in unicodedata.normalize("NFD", value)
        if unicodedata.category(c) != "Mn"
    ).lower()
    mots = [m for m in "".join(c if c.isalnum() else " " for c in plat).split()]
    return "-".join(mots)


def split_by_category(
    cards: list[dict[str, object]],
) -> dict[str, list[dict[str, object]]]:
    """Regroupe les cartes par catégorie, dans l'ordre d'apparition."""
    groupes: dict[str, list[dict[str, object]]] = {}
    for card in cards:
        categorie = str(card.get("category", ""))
        sans = {k: v for k, v in card.items() if k != "category"}
        groupes.setdefault(categorie, []).append(sans)
    return groupes


def _letters_only(value: str) -> str:
    """Ne garde que les lettres, sans accent, en minuscules.

    Jette donc les pastilles de couleur et la ponctuation dont une feuille
    décore ses niveaux : « 🟢 Facile » devient « facile ».
    """
    sans_accent = "".join(
        c
        for c in unicodedata.normalize("NFD", value)
        if unicodedata.category(c) != "Mn"
    )
    return "".join(c for c in sans_accent if c.isalpha()).lower()


def _parse_difficulty(cell: str, line: int, problems: list[str]) -> int:
    if cell == "":
        return DEFAULT_DIFFICULTY

    mot = _letters_only(cell)
    if mot in DIFFICULTY_ALIASES:
        return DIFFICULTY_ALIASES[mot]

    try:
        valeur = int(cell)
    except ValueError:
        niveaux = ", ".join(sorted(DIFFICULTY_ALIASES))
        problems.append(
            f"ligne {line} : difficulté « {cell} » illisible, attendu 1, 2 ou "
            f"3, ou l'un de : {niveaux}"
        )
        return DEFAULT_DIFFICULTY

    if valeur not in (1, 2, 3):
        problems.append(
            f"ligne {line} : difficulté {valeur} hors bornes, attendu 1, 2 ou 3"
        )
        return DEFAULT_DIFFICULTY

    return valeur


#: Colonnes du CSV que produit un classeur, dans cet ordre.
WORKBOOK_HEADER = ["texte", "difficulte", "categorie"]


def workbook_to_csv(path: Path) -> str:
    """Met un classeur à plat en CSV, **une feuille par catégorie**.

    La livraison arrive en `.xlsx`, un onglet par catégorie — vingt et un pour
    la livraison d'août 2026. Exporter chaque onglet à la main serait vingt et
    une occasions de se tromper de séparateur ou d'oublier une feuille.

    Le classeur ne porte **pas d'en-tête** : la première ligne est déjà une
    carte. On le sait parce qu'on lit une feuille et non un fichier plat — la
    position des colonnes y est stable — alors qu'un CSV anonyme ne permet pas
    de deviner si sa première ligne est un titre ou une donnée. D'où l'en-tête
    ajouté ici, qui rend la suite du chemin identique à celui d'un CSV.

    Rien n'est validé ici : longueur, doublons et niveaux illisibles restent
    l'affaire de [parse_csv], pour qu'il n'existe qu'un seul jeu de règles.
    """
    try:
        import openpyxl  # noqa: PLC0415 — dépendance du seul mode classeur
    except ModuleNotFoundError as absente:  # pragma: no cover
        raise ImportError_(
            "Lire un classeur demande openpyxl : pip install openpyxl"
        ) from absente

    classeur = openpyxl.load_workbook(path, data_only=True)
    sortie = io.StringIO()
    writer = csv.writer(sortie, delimiter=";", lineterminator="\n")
    writer.writerow(WORKBOOK_HEADER)

    for nom in classeur.sheetnames:
        feuille = classeur[nom]
        for row in feuille.iter_rows(values_only=True):
            cellules = ["" if c is None else str(c) for c in row]
            # Une feuille traîne des colonnes de travail vides sur toute sa
            # largeur : n'en garder que deux évite de les prendre pour des
            # champs. Les lignes de mise en page, elles, sont sautées ici
            # plutôt que signalées comme des cartes sans catégorie.
            if not any(c.strip() for c in cellules):
                continue
            texte = cellules[0] if cellules else ""
            niveau = cellules[1] if len(cellules) > 1 else ""
            writer.writerow([texte, niveau, nom])

    classeur.close()
    return sortie.getvalue()


def parse_csv(content: str) -> list[dict[str, object]]:
    """Lit une livraison et rend les cartes, ou lève en listant les fautes."""
    content = content.lstrip("﻿")
    if not content.strip():
        raise ImportError_("Livraison vide : aucune ligne à lire.")

    premiere = content.splitlines()[0]
    reader = csv.DictReader(
        io.StringIO(content), delimiter=_sniff_delimiter(premiere)
    )

    colonnes = _map_columns(list(reader.fieldnames or []))
    if "text" not in colonnes:
        attendus = ", ".join(sorted(HEADER_ALIASES["text"]))
        raise ImportError_(
            "Aucune colonne de texte trouvée. En-têtes acceptés : "
            f"{attendus}. En-têtes lus : {reader.fieldnames}"
        )

    cards: list[dict[str, object]] = []
    problems: list[str] = []
    vus: dict[str, int] = {}

    for index, row in enumerate(reader, start=2):
        texte = _clean_cell(row.get(colonnes["text"]))

        # Une ligne entièrement vide est de la mise en page, pas une carte
        # perdue : la signaler noierait les vraies fautes.
        if not any(_clean_cell(v) for v in row.values()):
            continue

        if not texte:
            problems.append(f"ligne {index} : texte vide")
            continue

        if len(texte) > MAX_TEXT_LENGTH:
            problems.append(
                f"ligne {index} : « {texte[:30]}… » fait {len(texte)} "
                f"caractères, {MAX_TEXT_LENGTH} au maximum"
            )
            continue

        categorie = ""
        if "category" in colonnes:
            categorie = _clean_cell(row.get(colonnes["category"]))
            # Une ligne sans catégorie, dans une feuille qui en a une, n'irait
            # nulle part : elle serait perdue en silence.
            if not categorie:
                problems.append(f"ligne {index} : catégorie vide")
                continue

        # Le doublon se juge sur **toute la livraison**, catégories confondues.
        # Deux catégories qui portent la même entrée ne la feraient sortir
        # qu'une fois au tirage (R6.4), mais le total annoncé au joueur avant
        # de démarrer (R6.2) compterait les deux : il jouerait avec une carte
        # de moins que ce qu'on lui a promis. C'est au contenu de trancher dans
        # quelle catégorie l'entrée vit.
        if texte in vus:
            problems.append(
                f"ligne {index} : « {texte} » déjà livré ligne {vus[texte]}"
            )
            continue
        vus[texte] = index

        difficulte = _parse_difficulty(
            _clean_cell(row.get(colonnes.get("difficulty", ""))),
            index,
            problems,
        )
        carte: dict[str, object] = {"text": texte, "difficulty": difficulte}
        if categorie:
            carte["category"] = categorie

        cards.append(carte)

    if problems:
        raise ImportError_(
            f"{len(problems)} ligne(s) inexploitable(s), rien n'a été écrit :\n"
            + "\n".join(f"  - {p}" for p in problems)
        )

    if not cards:
        raise ImportError_("Aucune carte lue : la livraison ne contient rien.")

    return cards


def build_deck(
    cards: list[dict[str, object]],
    *,
    existing: dict[str, object] | None,
    overrides: dict[str, object],
) -> dict[str, object]:
    """Assemble le deck, en reprenant les métadonnées déjà en place.

    `contentVersion` s'incrémente à chaque réécriture : c'est lui qui déclenche
    le re-seeding des installations existantes. L'oublier livrerait des cartes
    que personne ne verrait jamais.
    """
    deck: dict[str, object] = {
        "id": "",
        "name": "",
        "description": None,
        "icon": None,
        "audience": "family",
        "minAge": 6,
        "contentVersion": 0,
        "isPremium": False,
        "sortOrder": 0,
    }
    deck.update(existing or {})
    deck.update({k: v for k, v in overrides.items() if v is not None})

    deck["contentVersion"] = int(deck["contentVersion"]) + 1
    deck["cards"] = cards
    return {k: v for k, v in deck.items() if v is not None}


def _write_deck(
    path: Path,
    cards: list[dict[str, object]],
    overrides: dict[str, object],
) -> dict[str, object]:
    """Écrit un deck, en reprenant les métadonnées déjà en place."""
    existing = None
    if path.exists():
        existing = json.loads(path.read_text(encoding="utf-8"))
        existing.pop("cards", None)

    deck = build_deck(cards, existing=existing, overrides=overrides)
    path.write_text(
        json.dumps(deck, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return deck


def _write_many(
    args: argparse.Namespace,
    groupes: dict[str, list[dict[str, object]]],
) -> int:
    """Écrit un fichier par catégorie.

    L'identifiant vient du nom de la catégorie : il doit rester stable d'une
    livraison à l'autre, sinon le seeder croit à une catégorie neuve et
    l'ancienne reste en base à côté. Renommer une catégorie dans la feuille
    revient donc à en créer une autre — c'est signalé en fin de course.
    """
    args.out_dir.mkdir(parents=True, exist_ok=True)

    for nom, cards in groupes.items():
        identifiant = slugify(nom)
        chemin = args.out_dir / f"{identifiant}.json"
        connu = chemin.exists()

        deck = _write_deck(
            chemin,
            cards,
            {
                "id": identifiant,
                "name": nom,
                "audience": args.audience,
                "minAge": args.min_age,
                "sortOrder": args.sort_order,
                "isPremium": args.premium,
            },
        )

        etat = "mise à jour" if connu else "NOUVELLE"
        print(
            f"{len(cards):4d} cartes → {chemin.name} "
            f"(contentVersion {deck['contentVersion']}, {etat})"
        )

    print(
        "\nVérifie les catégories marquées NOUVELLE : un nom modifié dans la "
        "feuille produit un identifiant neuf, et l'ancien deck reste en base."
    )
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "csv",
        type=Path,
        help="livraison à convertir : CSV, TSV, ou classeur .xlsx",
    )
    parser.add_argument("--out", type=Path, help="JSON produit, une catégorie")
    parser.add_argument(
        "--out-dir",
        type=Path,
        help="dossier de sortie, une livraison à colonne « catégorie »",
    )
    parser.add_argument("--id")
    parser.add_argument("--name")
    parser.add_argument("--description")
    parser.add_argument("--icon")
    parser.add_argument("--audience", choices=["family", "adult"])
    parser.add_argument("--min-age", type=int, choices=[6, 10, 13, 18])
    parser.add_argument("--sort-order", type=int)
    parser.add_argument("--premium", action="store_true", default=None)
    args = parser.parse_args(argv)

    try:
        # Un classeur passe par le même chemin qu'un CSV : il est mis à plat,
        # puis relu par `parse_csv`. Les règles de refus n'existent qu'une
        # fois, quel que soit le format d'arrivée.
        if args.csv.suffix.lower() in {".xlsx", ".xlsm"}:
            cards = parse_csv(workbook_to_csv(args.csv))
        else:
            cards = parse_csv(args.csv.read_text(encoding="utf-8-sig"))
    except ImportError_ as erreur:
        print(f"Import refusé.\n{erreur}", file=sys.stderr)
        return 1
    except OSError as erreur:
        print(f"Livraison illisible : {erreur}", file=sys.stderr)
        return 1

    par_categorie = any("category" in c for c in cards)

    if par_categorie and not args.out_dir:
        print(
            "Cette livraison porte une colonne « catégorie » : utilise "
            "--out-dir pour écrire un fichier par catégorie.",
            file=sys.stderr,
        )
        return 1
    if not par_categorie and not args.out:
        print(
            "Livraison sans colonne « catégorie » : --out est requis.",
            file=sys.stderr,
        )
        return 1

    if par_categorie:
        return _write_many(args, split_by_category(cards))

    existing = None
    if args.out.exists():
        existing = json.loads(args.out.read_text(encoding="utf-8"))
        existing.pop("cards", None)

    deck = build_deck(
        cards,
        existing=existing,
        overrides={
            "id": args.id,
            "name": args.name,
            "description": args.description,
            "icon": args.icon,
            "audience": args.audience,
            "minAge": args.min_age,
            "sortOrder": args.sort_order,
            "isPremium": args.premium,
        },
    )

    if not deck["id"] or not deck["name"]:
        print(
            "Un deck neuf demande au moins --id et --name.", file=sys.stderr
        )
        return 1

    args.out.write_text(
        json.dumps(deck, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(
        f"{len(cards)} cartes écrites dans {args.out} "
        f"(contentVersion {deck['contentVersion']})"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
